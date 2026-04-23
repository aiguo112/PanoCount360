
import os
import sys
import json
import math
import argparse
import random
from pathlib import Path

PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.append(PROJECT_ROOT)

import numpy as np
import torch
from torch.utils.data import DataLoader
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

from data.panocount_dataset import PanoCountDataset
from models.model_factory import build_model


MODEL_LIST = ["csrnet", "mcnn", "can", "csrnet_pano"]
DENSITY_GROUPS = [
    ("Sparse", 0, 50),
    ("Medium", 50, 200),
    ("Dense", 200, 500),
    ("Ultra-dense", 500, float("inf")),
]


def safe_div(a, b):
    return a / b if b != 0 else 0.0


def summarize_metrics(preds, gts):
    preds = np.asarray(preds, dtype=np.float64)
    gts = np.asarray(gts, dtype=np.float64)
    err = preds - gts
    abs_err = np.abs(err)
    sq_err = err ** 2

    mae = float(abs_err.mean())
    rmse = float(np.sqrt(sq_err.mean()))
    mape = float(np.mean(abs_err / np.maximum(gts, 1.0)) * 100.0)
    smape = float(np.mean(2.0 * abs_err / np.maximum(np.abs(preds) + np.abs(gts), 1.0)) * 100.0)
    medae = float(np.median(abs_err))
    nae = float(np.sum(abs_err) / np.maximum(np.sum(gts), 1.0))
    bias = float(np.mean(err))
    rel_bias = float(np.sum(err) / np.maximum(np.sum(gts), 1.0))
    corr = float(np.corrcoef(preds, gts)[0, 1]) if len(preds) > 1 and np.std(preds) > 1e-9 and np.std(gts) > 1e-9 else 0.0

    ss_res = float(np.sum((gts - preds) ** 2))
    ss_tot = float(np.sum((gts - gts.mean()) ** 2))
    r2 = float(1.0 - ss_res / ss_tot) if ss_tot > 1e-9 else 0.0

    return {
        "MAE": mae,
        "RMSE": rmse,
        "MAPE_%": mape,
        "sMAPE_%": smape,
        "MedAE": medae,
        "NAE": nae,
        "Bias": bias,
        "Relative_Bias": rel_bias,
        "Pearson_r": corr,
        "R2": r2,
        "Num_Samples": int(len(gts)),
        "GT_Mean": float(gts.mean()) if len(gts) else 0.0,
        "Pred_Mean": float(preds.mean()) if len(preds) else 0.0,
    }


def get_density_group(gt_count):
    for name, lo, hi in DENSITY_GROUPS:
        if lo <= gt_count < hi:
            return name
    return "Unknown"


@torch.no_grad()
def run_model(model, loader, device):
    model.eval()
    rows = []
    for batch in loader:
        images = batch["image"].to(device, non_blocking=True)
        gts = batch["density"].to(device, non_blocking=True)
        ids = batch["id"]

        pred_density = model(images)
        pred_counts = pred_density.sum(dim=(1, 2, 3)).detach().cpu().numpy()
        gt_counts = gts.sum(dim=(1, 2, 3)).detach().cpu().numpy()

        for i in range(len(ids)):
            rows.append({
                "id": str(ids[i]),
                "gt_count": float(gt_counts[i]),
                "pred_count": float(pred_counts[i]),
                "error": float(pred_counts[i] - gt_counts[i]),
                "abs_error": float(abs(pred_counts[i] - gt_counts[i])),
                "sq_error": float((pred_counts[i] - gt_counts[i]) ** 2),
                "density_group": get_density_group(float(gt_counts[i])),
            })
    return rows


def add_sheet_title(ws, title, subtitle=None):
    ws["A1"] = title
    ws["A1"].font = Font(size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="1F4E78")
    ws["A1"].alignment = Alignment(horizontal="left")
    if subtitle:
        ws["A2"] = subtitle
        ws["A2"].font = Font(size=10, italic=True, color="555555")


def style_header_row(ws, row_idx, ncols):
    fill = PatternFill("solid", fgColor="D9EAF7")
    thin = Side(style="thin", color="B7C9D6")
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row_idx, column=c)
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.border = Border(bottom=thin)
        cell.alignment = Alignment(horizontal="center")


def autofit_width(ws, extra=2):
    max_width = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            l = len(str(cell.value))
            col = cell.column
            max_width[col] = max(max_width.get(col, 0), l)
    for col_idx, width in max_width.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = min(width + extra, 40)


def write_table(ws, start_row, headers, data_rows):
    for c, h in enumerate(headers, start=1):
        ws.cell(row=start_row, column=c, value=h)
    style_header_row(ws, start_row, len(headers))
    r = start_row + 1
    for row in data_rows:
        for c, v in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=v)
        r += 1
    return r - 1


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--split", type=str, default="test", choices=["train", "val", "test"])
    parser.add_argument("--image-height", type=int, default=512)
    parser.add_argument("--image-width", type=int, default=1024)
    parser.add_argument("--batch-size", type=int, default=2)
    parser.add_argument("--num-workers", type=int, default=0)
    parser.add_argument("--output-dir", type=str, default=None)
    args = parser.parse_args()

    project_root = Path(PROJECT_ROOT)
    dataset_root = project_root / "data" / "dataset"
    split_file = dataset_root / "splits" / f"{args.split}.txt"
    output_dir = Path(args.output_dir) if args.output_dir else project_root / "analysis_outputs" / args.split
    output_dir.mkdir(parents=True, exist_ok=True)

    device = "cuda" if torch.cuda.is_available() else "cpu"

    dataset = PanoCountDataset(
        dataset_root=str(dataset_root),
        split_file=str(split_file),
        image_size=(args.image_height, args.image_width),
        training=False,
    )
    loader = DataLoader(dataset, batch_size=args.batch_size, shuffle=False, num_workers=args.num_workers, pin_memory=True)

    per_model_rows = {}
    summary_rows = []
    per_bin_rows = []

    for model_name in MODEL_LIST:
        ckpt_path = project_root / "checkpoints" / model_name / "best_model.pth"
        if not ckpt_path.exists():
            print(f"[WARN] Missing checkpoint for {model_name}: {ckpt_path}")
            continue

        model = build_model(model_name).to(device)
        ckpt = torch.load(str(ckpt_path), map_location=device)
        model.load_state_dict(ckpt["model_state"])

        rows = run_model(model, loader, device)
        per_model_rows[model_name] = rows

        preds = [r["pred_count"] for r in rows]
        gts = [r["gt_count"] for r in rows]
        metrics = summarize_metrics(preds, gts)
        summary_rows.append([model_name] + [metrics[k] for k in [
            "Num_Samples", "MAE", "RMSE", "MAPE_%", "sMAPE_%", "MedAE", "NAE", "Bias", "Relative_Bias", "Pearson_r", "R2", "GT_Mean", "Pred_Mean"
        ]])

        for group_name, lo, hi in DENSITY_GROUPS:
            g_rows = [r for r in rows if lo <= r["gt_count"] < hi]
            if len(g_rows) == 0:
                continue
            g_preds = [r["pred_count"] for r in g_rows]
            g_gts = [r["gt_count"] for r in g_rows]
            g_metrics = summarize_metrics(g_preds, g_gts)
            per_bin_rows.append([
                model_name, group_name, f"{lo}-{hi if math.isfinite(hi) else '+'}",
                g_metrics["Num_Samples"], g_metrics["MAE"], g_metrics["RMSE"],
                g_metrics["MAPE_%"], g_metrics["sMAPE_%"], g_metrics["MedAE"],
                g_metrics["NAE"], g_metrics["Bias"], g_metrics["Pearson_r"], g_metrics["R2"]
            ])

    # Merge per-image table
    id_to_row = {}
    all_ids = set()
    for model_name, rows in per_model_rows.items():
        for r in rows:
            all_ids.add(r["id"])
            id_to_row.setdefault(r["id"], {"id": r["id"], "gt_count": r["gt_count"], "density_group": r["density_group"]})
            id_to_row[r["id"]][f"{model_name}_pred"] = r["pred_count"]
            id_to_row[r["id"]][f"{model_name}_error"] = r["error"]
            id_to_row[r["id"]][f"{model_name}_abs_error"] = r["abs_error"]

    per_image_rows = []
    for image_id in sorted(all_ids, key=lambda x: int(x) if str(x).isdigit() else str(x)):
        base = id_to_row[image_id]
        row = [base["id"], base["gt_count"], base["density_group"]]
        for model_name in MODEL_LIST:
            row.extend([
                base.get(f"{model_name}_pred", None),
                base.get(f"{model_name}_error", None),
                base.get(f"{model_name}_abs_error", None),
            ])
        per_image_rows.append(row)

    # Worst cases by model
    worst_rows = []
    for model_name, rows in per_model_rows.items():
        top = sorted(rows, key=lambda x: x["abs_error"], reverse=True)[:20]
        for r in top:
            worst_rows.append([
                model_name, r["id"], r["gt_count"], r["pred_count"], r["error"], r["abs_error"], r["density_group"]
            ])

    # Write JSON too
    with open(output_dir / f"detailed_eval_{args.split}.json", "w", encoding="utf-8") as f:
        json.dump({
            "summary": summary_rows,
            "per_bin": per_bin_rows,
            "per_image": per_image_rows,
            "worst_cases": worst_rows,
        }, f, indent=2)

    # Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    add_sheet_title(ws, f"PanoCount Detailed Evaluation ({args.split})",
                    "Models: CSRNet, MCNN, CAN, CSRNetPano | Metrics saved by split, density group, and per-image results")

    summary_headers = ["Model", "Num_Samples", "MAE", "RMSE", "MAPE_%", "sMAPE_%", "MedAE", "NAE", "Bias", "Relative_Bias", "Pearson_r", "R2", "GT_Mean", "Pred_Mean"]
    write_table(ws, 4, summary_headers, summary_rows)
    autofit_width(ws)

    ws2 = wb.create_sheet("Per_Density_Group")
    add_sheet_title(ws2, "Error vs Crowd Density", "Requested groups: Sparse 0-50, Medium 50-200, Dense 200-500, Ultra-dense 500+")
    per_bin_headers = ["Model", "Density_Group", "Count_Range", "Num_Samples", "MAE", "RMSE", "MAPE_%", "sMAPE_%", "MedAE", "NAE", "Bias", "Pearson_r", "R2"]
    write_table(ws2, 4, per_bin_headers, per_bin_rows)
    autofit_width(ws2)

    ws3 = wb.create_sheet("Per_Image")
    add_sheet_title(ws3, f"Per-Image Predictions ({args.split})")
    per_image_headers = ["Image_ID", "GT_Count", "Density_Group"]
    for model_name in MODEL_LIST:
        label = model_name.upper().replace("_", " ")
        per_image_headers.extend([f"{label}_Pred", f"{label}_Error", f"{label}_AbsError"])
    write_table(ws3, 4, per_image_headers, per_image_rows)
    autofit_width(ws3)

    ws4 = wb.create_sheet("Worst_Cases")
    add_sheet_title(ws4, "Top 20 Worst Cases Per Model")
    worst_headers = ["Model", "Image_ID", "GT_Count", "Pred_Count", "Error", "Abs_Error", "Density_Group"]
    write_table(ws4, 4, worst_headers, worst_rows)
    autofit_width(ws4)

    ws5 = wb.create_sheet("Recommendations")
    add_sheet_title(ws5, "Recommendations Before Designing the New Method")
    recommendations = [
        ["1", "Use CSRNet as the main baseline", "It is clearly the strongest baseline and should be the anchor for the new method."],
        ["2", "Analyze error vs crowd density", "Focus on Sparse (0–50), Medium (50–200), Dense (200–500), and Ultra-dense (500+) cases."],
        ["3", "Inspect 10 qualitative samples", "Look for failure modes such as missed dense regions, over-smoothing, and boundary artifacts."],
        ["4", "Prioritize ERP-aware improvements", "Most likely gains come from circular padding, latitude-aware weighting, and ERP-specific context modeling."],
        ["5", "Compare qualitative failures by density group", "Choose examples from sparse, medium, dense, and ultra-dense scenes to guide model design."],
    ]
    write_table(ws5, 4, ["#", "Recommendation", "Why"], recommendations)
    autofit_width(ws5)

    xlsx_path = output_dir / f"panocount_detailed_eval_{args.split}.xlsx"
    wb.save(str(xlsx_path))

    print(f"Saved detailed workbook to: {xlsx_path}")
    print(f"Saved JSON to: {output_dir / f'detailed_eval_{args.split}.json'}")


if __name__ == "__main__":
    main()
